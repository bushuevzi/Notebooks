#!/usr/share/anaconda3/bin/python3.5
# -*- coding: utf-8 -*-
"""
Created on Wed May  4 00:56:42 2016

@author: root
"""
########################
# import block
########################
import openpyxl

########################
# def block
########################


########################
# main block
########################
def main():
	# Берем значение картриджа из унифицированной формы
	unidb_wb = openpyxl.load_workbook('.\\unidb_wb.xlsx')
	unidb_rashodniki = unidb_wb.get_sheet_by_name('РАСХОДКА')
	#print(unidb_rashodniki['A1'].value)
	# экспорт значений картриджей в словарь {partNumber: (manufacture, description)}
	unidb_kartridg = {}
	for row in unidb_rashodniki['A2':'E1076']:
		#print(row[1].value, row[2].value, row[3].value)
		unidb_kartridg[row[2].value] = (row[1].value, row[3].value)
	#print(unidb_kartridg)
	unidb_zip = unidb_wb.get_sheet_by_name('ЗИП')
	for row in unidb_zip['A2':'E244']:
		unidb_kartridg[row[2].value] = (row[1].value, row[3].value)
	
	#Создадим файл как в шаблон потребности из москвы
	output_wb = openpyxl.load_workbook('.\\Приложение №1_шаблон.xlsx')
	output_sheet = output_wb.active
	#output_list = []
	output_list = []
		

	# Ищем картридж в портебности
	name_file_potrebnost = 'potrebnost_UNP_CO.xlsx'
	path_to_potrebnost = '.\\' + name_file_potrebnost
	
	
	potrebnost_wb = openpyxl.load_workbook(path_to_potrebnost)
	potrebnost_sheet = potrebnost_wb.active
	zakazchik = potrebnost_sheet.cell(column = 4, row = 4).value
	adress = potrebnost_sheet.cell(column = 5, row = 4).value
	for partNumber in unidb_kartridg.keys():
		amount = 0
		for row in potrebnost_sheet['A4':'F200']:	
			if str(row[5].value).lower() == 'новый':
				if str(partNumber) in row[1].value:
					print(partNumber)
					#while str(partNumber) in row[1].value:
					amount = amount + int(row[2].value)
					for i in row:
						i.value = ""
						#print(i.value)
		print(amount)
		if amount > 0:
			output_list.append((partNumber, unidb_kartridg[partNumber][1],unidb_kartridg[partNumber][0], amount))				
	# ОСТАТКИ
	path_to_potrebnost_after_script = '.\\after_script_' + name_file_potrebnost
#	# убираем пустые строки
#	for row_of_cell in potrebnost_sheet['A4':'F200']:
#		if row_of_cell[1].value != "":
#			for i in range(6):
#				potrebnost_sheet.cell(column = i+1, row = potrebnost_sheet['A4':'F200'].index(row_of_cell)+1).value = row_of_cell[i].value	
		
	potrebnost_wb.save(path_to_potrebnost_after_script)
	
	# Если картридж новый то переносим данную позицию из потребности в форму
	number_position = 0
	for record in output_list:
		output_sheet.cell(column = 1, row = 8+number_position).value = number_position+1 # нумеруем позицию
		for i in range(4):
			output_sheet.cell(column = i+2, row = 8+number_position).value = record[i]
		output_sheet.cell(column = 6, row = 8+number_position).value = 'шт.' # ставим единицу измерения.
		number_position = number_position+1
	# прописываем наименование заказчика и адрес поставки под таблицей
	output_sheet.cell(column = 1, row = 86).value = "Грузополучатель  (конечный пользователь): " + zakazchik   # наименование заказчика
	output_sheet.cell(column = 1, row = 88).value = "Адрес доставки: " + adress  # Адрес поставки	
	path_to_output_file = '.\\Приложение №1_' + name_file_potrebnost	
	output_wb.save(path_to_output_file)
	

if __name__ == '__main__':
    main()
