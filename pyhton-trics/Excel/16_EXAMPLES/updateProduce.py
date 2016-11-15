#!/usr/share/anaconda3/bin/python3.5
# -*- coding: utf-8 -*-
"""
Created on Wed May  4 00:56:42 2016

@author: root
"""
########################
# import block
########################

import openpyxl

########################
# def block
########################

########################
# main block
########################
def main():
    wb = openpyxl.load_workbook('../data/automate_online-materials/produceSales.xlsx')
    sheet = wb.get_sheet_by_name('Sheet')

    # The produce type and thei updated prices
    PRICE_UPDATES = {'Garlic': 3.07,
                    'Celery': 1.19,
                    'Lemon': 1.27}
    # Loop through the rows and update the prices.
    for rowNum in range(2, sheet.max_row):
        produceName = sheet.cell(row=rowNum, column=1).value
        if produceName in PRICE_UPDATES:
            sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
    wb.save('updatedProduceSales.xlsx')


if __name__ == '__main__':
    main()
